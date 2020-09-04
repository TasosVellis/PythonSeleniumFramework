import openpyxl
book = openpyxl.load_workbook("C:\\Users\\Angie\\PythonDemo.xlsx")
sheet = book.active
dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Tasos"

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

for i in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(2,sheet.max_column+1):  #to get columns
            #dict ["lastname']="Vellis
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value